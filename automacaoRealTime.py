#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Jun 30 19:39:30 2019

@author: alex
"""


from bs4 import BeautifulSoup
#import requests
import re
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from xlrd import open_workbook, XLRDError
from selenium import webdriver
from openpyxl import load_workbook
from os import listdir
import config
from selenium.webdriver.support.ui import Select
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
import time
import os
import os.path
import shutil
#from xhtml2pdf import pisa   
import json
import xlsxwriter
import sys
#reload(sys)
#sys.setdefaultencoding('utf-8')



def IniciarAutomacao():  
    
    #lista os arquivos de faixa a serem processados
    arquivos = listar_arquivos(caminho=config.PATH_CONFIG['resourcesPath'])

    #loop para pegar todos os arquivos de faixas
    for arq in arquivos:

        #Verifica se o arquivo é um xlsx
        if verificarExcel(config.PATH_CONFIG['resourcesPath'] + arq):

            nmArqSplit = arq.split('.')
            dataArquivo = nmArqSplit[0].split('_')
            #Extrai dia mes e ano do nome do arquivo
            diaArquivo = int(dataArquivo[0])
            mesArquivo = int(dataArquivo[1])
            anoArquivo = int(dataArquivo[2])

            #Cria uma pasta para salvar os arquivos dos relatórios
            nmDiretorioSalvarRelatorio = str(diaArquivo)+'_'+str(switch_demo(int(mesArquivo)))+'_'+str(anoArquivo)
            if os.path.exists(config.PATH_CONFIG['pathRelatorio'] + nmDiretorioSalvarRelatorio):
                shutil.rmtree(config.PATH_CONFIG['pathRelatorio'] + nmDiretorioSalvarRelatorio)
            else:
                os.mkdir('./'+nmDiretorioSalvarRelatorio, 0o777)
                pathRelatorio = config.PATH_CONFIG['pathRelatorio'] + nmDiretorioSalvarRelatorio


            appState ={
                "recentDestinations": [
                    {
                        "id": "Save as PDF",
                        "origin": "local"
                    }
                ],
                "selectedDestinationId": "Save as PDF",
                "version": 2
            }


            options = webdriver.ChromeOptions()

            options.add_experimental_option("prefs", {
             "download.default_directory": config.PATH_CONFIG['pathRelatorio'] + nmDiretorioSalvarRelatorio,
             "download.prompt_for_download": False,
             "download.directory_upgrade": True,
             "safebrowsing_for_trusted_sources_enabled": False,
             "safebrowsing.enabled": False,
             'safebrowsing.disable_download_protection': True,
             "multiple-automatic-downloads" : 1,
             'savefile.default_directory':config.PATH_CONFIG['pathRelatorio'] + nmDiretorioSalvarRelatorio,
             'printing.print_preview_sticky_settings.appState': {"recentDestinations": [ {"id": "Save as PDF","origin": "local" } ], "selectedDestinationId": "Save as PDF", "version": 2 }
            })

            options.add_argument('--kiosk-printing')


            #Função para omitir o navegador
            #options.add_argument('--headless')
            driver = webdriver.Chrome(executable_path=config.PATH_CONFIG['webDrivePathChrome'], chrome_options=options)

            username = config.REAL_TIME_CONFIG['user']
            password =  config.REAL_TIME_CONFIG['password']
            getdriver = config.REAL_TIME_CONFIG['url']
            driver.get(getdriver)

            element = WebDriverWait(driver, 3000).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="main"]/div/div/div/div/form/fieldset[2]/div/div[1]/label/span'))
            )
            #Seta login e senha
            driver.find_element_by_name('username').send_keys(username)
            driver.find_element_by_name('password').send_keys(password)

            #pega todo html da pagina para extrair a pergunta do captcha
            soup = BeautifulSoup(driver.page_source, 'lxml')

            # Pegar todo o texto da div legend e pega a pergunta do captcha
            legend = soup.find_all('legend',text=True)
            perguntaSplit  = legend[1].text.split(' ')
            soma = int(perguntaSplit[2]) + int(perguntaSplit[4])

            #Escolhe o a resposta e marca o termos de uso e entra
            driver.find_element_by_xpath('//span[text()="'+str(soma)+'"]').click()
            driver.find_element_by_xpath('//*[@id="main"]/div/div/div/div/form/fieldset[3]/div/div/label/i').click()
            driver.find_element_by_xpath('//*[@id="main"]/div/div/div/div/form/div/div/input').click()

            #aguarda ate que a pagina seja carregada
            element = WebDriverWait(driver, 3000).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="main"]/div/div/menu/div[3]/button[2]/span[1]'))
            )


            indexArrOutPut = 0
            arrOutput = []
            arr = ['FAIXA', 'PROGRAMA', 'ARATU', 'RECORD TV ITAPOAN', 'TV BAND', 'BAHIA', 'TOTAL LIGADOS', 'REDE TV!', 'TVE BAHIA', 'REDE BRASIL', 'TOTAL LIGADOS ESPECIAL']
            arrOutput.append(arr)

            escolherData(driver,arq)


            #Abre o menu
            driver.find_element_by_xpath("//button[contains(@type, 'menuToggle')]").click()



            arrFaixasTratadas = obterFaixas(config.PATH_CONFIG['resourcesPath'] + arq)

            for faixa in arrFaixasTratadas:
                arr = []
                arr.append(str(faixa[0]))
                arr.append(str(faixa[1]))

                horaInicial = str(faixa[2]).replace("1899-12-30 ", "").split(":")
                horaFinal = str(faixa[3]).replace("1899-12-30 ", "").split(":")


                driver.find_element_by_xpath("//*[@id='menu-container']/div/div/div/div[6]/ol/li[1]/a").click()
               # time.sleep(2)

                element = WebDriverWait(driver, 3000).until(
                    EC.presence_of_element_located(
                        (By.XPATH, '//h1[text()="Enviar para"]'))
                )




                selectHoraInicial = Select(driver.find_element_by_xpath(
                    '//*[@id="menu-container"]/div/div[2]/div/div[1]/div[1]/div/div[1]/select'))
                selectMinutoInicial = Select(driver.find_element_by_xpath(
                    '//*[@id="menu-container"]/div/div[2]/div/div[1]/div[1]/div/div[2]/select'))
                selectHoraFinal = Select(driver.find_element_by_xpath(
                    '//*[@id="menu-container"]/div/div[2]/div/div[1]/div[2]/div/div[1]/select'))
                selectMinutoFinal = Select(driver.find_element_by_xpath(
                    '//*[@id="menu-container"]/div/div[2]/div/div[1]/div[2]/div/div[2]/select'))

                selectHoraInicial.select_by_value(horaInicial[0])
                selectMinutoInicial.select_by_value(horaInicial[1])

                selectHoraFinal.select_by_value(horaFinal[0])
                selectMinutoFinal.select_by_value(horaFinal[1])

                paginaPrincipal = driver.window_handles[0]

                driver.find_element_by_xpath('//*[@id="menu-container"]/div/div[2]/div/div[2]/ol/li/div/a').click()

                time.sleep(2)

                if len(driver.window_handles) == 1:
                    driver.find_element_by_xpath('//*[@id="main"]/div/div/div[8]/div/div/div/form/div/input').click()

                    time.sleep(1)

                else:

                    popup = driver.window_handles[1]

                    driver.switch_to.window(popup)

                    while AguardarElemento(driver, 'gridTable') == False:
                        AguardarElemento(driver, 'gridTable')

                    page_source = driver.page_source
                    dadosMinAMin = BeautifulSoup(page_source, 'lxml')

                    driver.find_element_by_id('exportCSV').click()
                    time.sleep(1)

                    #driver.execute_script('window.print();')

                    spans = []
                    tfoot = dadosMinAMin.find('tfoot')
                    for foot in tfoot:
                        spans = foot.find_all('span')

                    arr.append(str(spans[2].text))
                    arr.append(str(spans[3].text))
                    arr.append(str(spans[4].text))
                    arr.append(str(spans[5].text))
                    arr.append(str(spans[6].text))
                    arr.append(str(spans[7].text))
                    arr.append(str(spans[8].text))
                    arr.append(str(spans[9].text))
                    arr.append(str(spans[10].text))


                    arrOutput.append(arr)

                    driver.switch_to.window(paginaPrincipal)


            gerarOUTPUT(arrOutput, config.PATH_CONFIG['pathRelatorio'] + nmDiretorioSalvarRelatorio  + '/' + 'output_' + str(dataArquivo[0]) + '_' + str(dataArquivo[1]) + '_' +  str(dataArquivo[2]) +'.xlsx')
            driver.quit()

def gerarOUTPUT(dados, path):

    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(path)
    worksheet = workbook.add_worksheet()

    row = 0

    for linha in dados:
        col = 0
        for coluna in linha:
            worksheet.write(row, col, coluna)
            col += 1

        row += 1

    workbook.close()


def obterFaixas(path):
    # Abre a planilha
    book = load_workbook(path)
    sheet = book.active
    qtdLinhasPlaniha = sheet.max_row

    rangeHorarios = sheet['A2': 'D' + str(qtdLinhasPlaniha)]

    arrFaixasProgramas = []
    faixas = []
    programas = []


    horaIniQVP = None
    horaFimUniverso = None
    i = 1

    for faixa, programa, c1, c2 in rangeHorarios:

        if str(programa.value).strip() == "QUE VENHA O POVO":
            horaIniQVP = str(c1.value)

        if str(programa.value).strip() == "UNIVERSO":
            horaFimUniverso = str(c2.value)

    if (horaIniQVP is not None) and (horaFimUniverso is not None):
        arr = []
        arr.append(0)
        arr.append("Local_1")
        arr.append(horaIniQVP)
        arr.append(horaFimUniverso)
        faixas.append(arr)




    for faixa, programa, c1, c2 in rangeHorarios:
        arr = []

        if faixa.value != None:

            if str(programa.value).strip() == "Local_2":
                arr.append(0)
                arr.append(str(programa.value))
                arr.append(str(c1.value))
                arr.append(str(c2.value))
                faixas.append(arr)
            elif str(programa.value).strip() == "Geral_dia":
                arr.append(0)
                arr.append(str(programa.value))
                arr.append(str(c1.value))
                arr.append(str(c2.value))
                faixas.append(arr)
            elif str(programa.value).strip() == "Geral_24h":
                arr.append(0)
                arr.append(str(programa.value))
                arr.append(str(c1.value))
                arr.append(str(c2.value))
                faixas.append(arr)
            elif str(programa.value).strip() == "Matutino":
                arr.append(0)
                arr.append(str(programa.value))
                arr.append(str(c1.value))
                arr.append(str(c2.value))
                faixas.append(arr)
            elif str(programa.value).strip() == "Vespertino":
                arr.append(0)
                arr.append(str(programa.value))
                arr.append(str(c1.value))
                arr.append(str(c2.value))
                faixas.append(arr)
            elif str(programa.value).strip()  == "Noturno":
                arr.append(0)
                arr.append(str(programa.value))
                arr.append(str(c1.value))
                arr.append(str(c2.value))
                faixas.append(arr)
            elif str(programa.value).strip() == "Madrugada":
                arr.append(0)
                arr.append(str(programa.value))
                arr.append(str(c1.value))
                arr.append(str(c2.value))
                faixas.append(arr)
            elif str(programa.value).strip() == "Local_3":
                arr.append(0)
                arr.append(str(programa.value))
                arr.append(str(c1.value))
                arr.append(str(c2.value))
                faixas.append(arr)
            else:
                arr.append(str(faixa.value))
                arr.append(str(programa.value))
                arr.append(str(c1.value))
                arr.append(str(c2.value))
                programas.append(arr)
                i = i + 1


    for faixa in faixas:
        faixa[0] = i
        i = i + 1


    for programa in programas:
        arrFaixasProgramas.append(programa)

    for faixa in faixas:
        arrFaixasProgramas.append(faixa)



    return arrFaixasProgramas





def escolherData(driver, arq):


    nmArqSplit = arq.split('.')
    dataArquivo = nmArqSplit[0].split('_')
    #Extrai dia mes e ano do nome do arquivo
    diaArquivo = int(dataArquivo[0])
    mesArquivo = int(dataArquivo[1])
    anoArquivo = int(dataArquivo[2])


    #Abre o calendário
    driver.find_element_by_xpath('//*[@id="main"]/div/div/menu/div[2]/button/span').click()

    #Escolhe uma data qualquer so pra fazer o calendario funcioinar
    driver.find_element_by_xpath("//*[@id='rowPanelDrawer']/div[1]/ol/div/div/div[2]/div[3]/div[1]/div[1]").click()
    time.sleep(3)


    #Verifica qual o mes e ano selecionado no calendario
    tituloCalendario = driver.find_element_by_xpath('//*[@id="rowPanelDrawer"]/div[1]/ol/div/div/div[2]/div[1]')
    mesAnoSistema  = tituloCalendario.text.split(" ");

      #Enquanto o mes e ano não for igual ao do arquivo, volta a data no calendario
    while(int(mesAnoSistema[1]) != anoArquivo):

        if(anoArquivo > int(mesAnoSistema[1])):
            driver.find_element_by_xpath("//*[@id='rowPanelDrawer']/div[1]/ol/div/div/div[1]/span[2]").click()
        else:
            driver.find_element_by_xpath("//*[@id='rowPanelDrawer']/div[1]/ol/div/div/div[1]/span[1]").click()

        tituloCalendario = driver.find_element_by_xpath('//*[@id="rowPanelDrawer"]/div[1]/ol/div/div/div[2]/div[1]')
        mesAnoSistema  = tituloCalendario.text.split(" ");

    while(str(mesAnoSistema[0]) != str(switch_demo(mesArquivo))):

        numMesSistema = obterNumeroMes(str(mesAnoSistema[0]))

        if(mesArquivo > int(numMesSistema)):
            driver.find_element_by_xpath("//*[@id='rowPanelDrawer']/div[1]/ol/div/div/div[1]/span[2]").click()
        else:
            driver.find_element_by_xpath("//*[@id='rowPanelDrawer']/div[1]/ol/div/div/div[1]/span[1]").click()


        tituloCalendario = driver.find_element_by_xpath('//*[@id="rowPanelDrawer"]/div[1]/ol/div/div/div[2]/div[1]')
        mesAnoSistema  = tituloCalendario.text.split(" ");




    #Ao chegar no mes e ano correto escolhe a data
    driver.find_element_by_xpath("//div[text()='"+str(diaArquivo)+"' and @class='DayPicker-Day']").click()
    time.sleep(3)




def verificarExistenciaElementoByXPATH(driver,xpath):
    try:
         WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, xpath)))
         return True
    except NoSuchElementException:
        return False


def AguardarElemento(driver, Id):

    try:
        WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.ID, Id)))
        return True
    except TimeoutException:
        return False


def VerifcarAlert(driver):

    try:
        WebDriverWait(driver, 1).until(EC.alert_is_present())
        return True
    except TimeoutException:
        return False



def listar_arquivos(caminho=None):
    lista_arqs = [arq for arq in listdir(caminho)]
    return lista_arqs

def verificarExcel(filename):
    try:
        open_workbook(filename)
    except XLRDError:
        return False
    else:
        return True

def switch_demo(argument):
    switcher = {
        1: "Janeiro",
        2: "Fevereiro",
        3: "Março",
        4: "Abril",
        5: "Maio",
        6: "Junho",
        7: "Julho",
        8: "Agosto",
        9: "Setembro",
        10: "Outubro",
        11: "Novembro",
        12: "Dezembro"
    }
    return switcher.get(argument, "Invalid month")

def obterNumeroMes (argument):
    switcher = {
       "Janeiro"  : 1,
       "Fevereiro": 2,
       "Março"    : 3,
       "Abril"    : 4,
       "Maio"     : 5,
       "Junho"    : 6,
       "Julho"    : 7,
       "Agosto"   : 8,
       "Setembro" : 9,
       "Outubro"  : 10,
       "Novembro" : 11,
       "Dezembro" : 12
    }
    return switcher.get(argument, "Invalid month")




IniciarAutomacao()
#obterFaixas(config.PATH_CONFIG['resourcesPath'] + "10_07_2019.xlsx")
